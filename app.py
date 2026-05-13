import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import io
import requests
from datetime import datetime

# --- હેલ્પર ફંક્શન્સ (આને સૌથી ઉપર રાખવા જરૂરી છે) ---
def fmt_date(val):
    if isinstance(val, datetime):
        return val.strftime("%d-%m-%Y")
    return ""

def blank_if_na_or_empty(val, is_date=False):
    if val is None:
        return ""
    if isinstance(val, str) and val.strip().upper() == "NA":
        return ""
    if isinstance(val, str) and val.strip() == "":
        return ""
    return fmt_date(val) if is_date else val

def center_cell(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "General"

def percent_text(val):
    if isinstance(val, (int, float)):
        return f"{int(val * 100)}%"
    return str(val)

def get_web_workbook(url):
    try:
        if "spreadsheets/d/" in url:
            file_id = url.split("spreadsheets/d/")[1].split("/")[0]
        elif "id=" in url:
            file_id = url.split("id=")[1].split("&")[0]
        else:
            st.error("❌ લિંક ફોર્મેટ ખોટું છે.")
            return None

        d_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
        response = requests.get(d_url)
        if response.status_code == 200:
            return load_workbook(filename=io.BytesIO(response.content), data_only=True)
        else:
            st.error("❌ ફાઇલ એક્સેસ નકારી. ગૂગલ ડ્રાઇવ શેરિંગ સેટિંગ્સ ચેક કરો.")
            return None
    except Exception as e:
        st.error(f"❌ એરર: {e}")
        return None

# --- લોગિન સિસ્ટમ ---
def check_password():
    if "password_correct" not in st.session_state:
        st.text_input("Please Enter Password", type="password", on_change=lambda: st.session_state.update({"password_correct": st.session_state.pwd == st.secrets["password"]}), key="pwd")
        return False
    return st.session_state["password_correct"]

if not check_password():
    st.stop()

# --- પેજ સેટઅપ ---
st.set_page_config(page_title="AIPL LOT Generator", layout="wide")
st.title("📊 Quality Data: RT/DPT LOT")

# --- સાઇડબાર ઇનપુટ્સ ---
with st.sidebar:
    st.header("⚙️ સેટિંગ્સ")
    # આ લાઇન સાઇડબારની બહાર લખવી જેથી તે દેખાય નહીં
    lhs_link = "https://docs.google.com/spreadsheets/d/1P1-U_1rhYJ28drrdGgwKBVntP9Uh4nlQ/edit?usp=drive_link&ouid=112560164772488108224&rtpof=true&sd=true"
    template_file = "LOT MASTER.xlsx"
    
    st.markdown("---")
    ndt_type = st.radio("🎯 Select NDT Type ", ["RT", "DPT"])
    lot_no = st.text_input("🔢 Please Enter LOT Number ")
    
    eb_joint = "N"
    if ndt_type == "DPT":
        eb_joint = st.radio("🛠️ Joint Type EB છે?", ["Y", "N"])

# --- મુખ્ય લોજિક ---
if st.button("🚀 રિપોર્ટ જનરેટ કરો"):
    if not lhs_link or not lot_no:
        st.error("⚠️ લિંક અને LOT નંબર જરૂરી છે.")
    else:
        with st.spinner("⏳ Please wait, processing your data..."):
            lhs_wb = get_web_workbook(lhs_link)
            
            if lhs_wb:
                if "Sheet2" not in lhs_wb.sheetnames:
                    st.error("❌ Sheet2 મળી નથી.")
                else:
                    lhs_ws = lhs_wb["Sheet2"]
                    filtered_rows = []
                    
                    if ndt_type == "RT":
                        for r in range(2, lhs_ws.max_row + 1):
                            if str(lhs_ws[f"AL{r}"].value).strip() == lot_no:
                                filtered_rows.append(r)
                        if filtered_rows:
                            rt_vals = set(percent_text(lhs_ws[f"AG{r}"].value) for r in filtered_rows if lhs_ws[f"AG{r}"].value)
                            wps_vals = set(str(lhs_ws[f"Y{r}"].value) for r in filtered_rows if lhs_ws[f"Y{r}"].value)
                            header_ndt = "RT-" + " & ".join(sorted(rt_vals))
                            header_wps = "WPS-" + " & ".join(sorted(wps_vals))
                            col_report, col_date = "AH", "AI"
                    else:
                        col_search = "T" if eb_joint == "Y" else "AF"
                        for r in range(2, lhs_ws.max_row + 1):
                            if str(lhs_ws[f"{col_search}{r}"].value).strip() == lot_no:
                                filtered_rows.append(r)
                        if filtered_rows:
                            if eb_joint == "Y":
                                header_ndt, header_wps = "DPT-10%", "WPS-NA"
                                col_report, col_date = "U", "V"
                            else:
                                dpt_vals = set(percent_text(lhs_ws[f"AC{r}"].value) for r in filtered_rows if lhs_ws[f"AC{r}"].value)
                                wps_vals = set(str(lhs_ws[f"Y{r}"].value) for r in filtered_rows if lhs_ws[f"Y{r}"].value)
                                header_ndt = "DPT-" + " & ".join(sorted(dpt_vals))
                                header_wps = "WPS-" + " & ".join(sorted(wps_vals))
                                col_report, col_date = "AD", "AE"

                    if not filtered_rows:
                        st.warning(f"⚠️ LOT {lot_no} માટે કોઈ ડેટા મળ્યો નથી.")
                    else:
                        try:
                            fmt_wb = load_workbook(template_file)
                            fmt_ws = fmt_wb.active
                            
                            fmt_ws["I4"].value = header_ndt
                            fmt_ws["D4"].value = header_wps
                            fmt_ws["N4"].value = f"LOT NAME: MPL/AIPL/{ndt_type}/LOT-{lot_no}"
                            fmt_ws["G7"].value = int(lot_no) if lot_no.isdigit() else lot_no
                            
                            row_out, sr, last_spool = 7, 1, None

                            for r in filtered_rows:
                                fmt_ws[f"A{row_out}"].value = sr
                                fmt_ws[f"B{row_out}"].value = lhs_ws[f"F{r}"].value
                                fmt_ws[f"D{row_out}"].value = lhs_ws[f"R{r}"].value
                                fmt_ws[f"E{row_out}"].value = lhs_ws[f"H{r}"].value
                                fmt_ws[f"F{row_out}"].value = f'{lhs_ws[f"K{r}"].value}"/{lhs_ws[f"P{r}"].value}'
                                fmt_ws[f"H{row_out}"].value = lhs_ws[f"X{r}"].value
                                fmt_ws[f"I{row_out}"].value = lhs_ws[f"J{r}"].value
                                fmt_ws[f"K{row_out}"].value = lhs_ws[f"AA{r}"].value
                                fmt_ws[f"L{row_out}"].value = fmt_date(lhs_ws[f"AB{r}"].value)
                                fmt_ws[f"M{row_out}"].value = "RT" if str(lhs_ws[f"J{r}"].value) == "BW" else "DPT"
                                
                                report_no = blank_if_na_or_empty(lhs_ws[f"{col_report}{r}"].value)
                                xr_no = blank_if_na_or_empty(lhs_ws[f"AJ{r}"].value)
                                fmt_ws[f"N{row_out}"].value = f"{report_no} ({xr_no})" if report_no and xr_no else report_no
                                
                                fmt_ws[f"O{row_out}"].value = fmt_date(lhs_ws[f"{col_date}{r}"].value)
                                fmt_ws[f"Q{row_out}"].value = fmt_date(lhs_ws[f"AR{r}"].value)

                                spool = lhs_ws[f"G{r}"].value
                                if spool != last_spool:
                                    fmt_ws[f"P{row_out}"].value = spool
                                    last_spool = spool

                                for col in "ABCDEFHIJKLMNOPQ":
                                    center_cell(fmt_ws[f"{col}{row_out}"])
                                row_out += 1
                                sr += 1

                            output = io.BytesIO()
                            fmt_wb.save(output)
                            output.seek(0)
                            
                            st.success(f"✅ તૈયાર! {len(filtered_rows)} રેકોર્ડ્સ.")
                            st.download_button(label="📥 એક્સેલ રિપોર્ટ ડાઉનલોડ કરો", data=output, file_name=f"{ndt_type}_LOT_{lot_no}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        except Exception as e:
                            st.error(f"❌ ટેમ્પલેટ લોડ કરવામાં ભૂલ: {e}")
